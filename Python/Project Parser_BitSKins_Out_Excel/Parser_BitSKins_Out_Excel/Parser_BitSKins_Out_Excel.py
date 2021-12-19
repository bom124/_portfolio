import math
import time
import pandas
import requests
import openpyxl
import json
import re
import socks
import socket


API_CODE='836da3b0-a1a4-401f-9bea-903b2e10de8b'
DOTA_CODE=570
CS_GO_CODE=730
DEFAULT_TOR_CHANNEL = 9150
EXCEL_FILE_NAME = "outPrice.xlsx"



def get_price_bitskins_api(code,game_code):
    res= requests.get('https://bitskins.com/api/v1/get_price_data_for_items_on_sale/?api_key=836da3b0-a1a4-401f-9bea-903b2e10de8b&code='+code+'&app_id='+str(game_code))
    data = res.json()
    data1 = data["data"]["items"]
    data2 = json.dumps(data1)
    return data2


def update_bitSkins_base_to_excel(game_code):
    print("Input autendefication-code:")
    code = input()
    data1 = get_price_bitskins_api(code,game_code)
    pandas.read_json(data1).to_excel(EXCEL_FILE_NAME)
    print("Sucesfull")

def request_tor(tor_channel):
    socks.set_default_proxy(socks.SOCKS5, "localhost", tor_channel)
    socket.socket = socks.socksocket

def get_accurate_price_steam(game_code,use_tor_proxy,tor_channel,count_items):
    if(use_tor_proxy):
     request_tor(tor_channel)
    d = {"company":1}
    start=0
    while(start<count_items):
        try:
            t=start
            rs = requests.get("https://steamcommunity.com/market/search/render/?sort_column=name&sort_dir=desc&appid="+str(game_code)+"&norender=1&count=100&start="+str(start))
            data = rs.json()
            print(start)
            for i in range(0,100):
                value1 = data["results"][i]["hash_name"]
                price = data["results"][i]["sell_price_text"]
                count = data["results"][i]["sell_listings"]
                d[value1]=price,count
            start+=100
        except:
            start=t
            print("error")
    return d

def save_steam_price_to_excele(d,count_items):
    file = openpyxl.load_workbook(EXCEL_FILE_NAME)
    sheet = file.get_sheet_by_name('Sheet1')
    sheet.cell(row=1, column=9).value = "Steam price"
    sheet.cell(row=1, column=10).value = "Steam count"
    for x in range(2,count_items+1):
        try:
            value1 = sheet.cell(row=x, column=2).value
            print(value1)
            sheet.cell(row=x, column=9).value = d[value1][0]
            print(d[value1][0])
            sheet.cell(row=x, column=10).value = d[value1][1]
        except:
            print("error")
    file.save(EXCEL_FILE_NAME)


def update_accurate_price_steam_to_excel(game_code,use_tor_proxy,tor_channel):
    file = openpyxl.load_workbook(EXCEL_FILE_NAME)
    sheet = file.get_sheet_by_name('Sheet1')
    data = get_accurate_price_steam(game_code,use_tor_proxy,tor_channel,sheet.max_row)
    print("Press any button to start saving data to a file")
    input()
    save_steam_price_to_excele(data,sheet.max_row)

def get_fast_price_steam_api(game_code):
    if(game_code==DOTA_CODE):
        rs = requests.get("http://dota2.csgobackpack.net/api/GetItemsList/v2/?no_details")
    else:
        rs = requests.get("http://csgobackpack.net/api/GetItemsList/v2/?no_details")
    data = rs.json()
    file = openpyxl.load_workbook(EXCEL_FILE_NAME)
    sheet = file.get_sheet_by_name('Sheet1')
    sheet.cell(row=1, column=9).value = "Steam price"
    for x in range(2,sheet.max_row+1):
        try:
            value1 = sheet.cell(row=x, column=2).value
            print(value1)
            sheet.cell(row=x, column=9).value = json.dumps(data["items_list"][value1]["price"]["24_hours"]["average"])
            print(data["items_list"][value1]["price"]["24_hours"]["lowest_price"])
        except:
            sheet.cell(row=x, column=9).value = sheet.cell(row=x, column=4).value
            print("error")
    file.save(EXCEL_FILE_NAME)

def program():
    while True:
     print("Choose a game:")
     print("1 - Dota2")
     print("2 - CS:GO")
     game_code = input()
     if(game_code=="1"):
         game_code = DOTA_CODE
     elif(game_code=="2"):
         game_code = CS_GO_CODE
     else:
         print("Wrong number")
         continue
     print("Update bitskins price?")
     print("1 - Yes")
     print("2 - No")
     approval = input()
     if(approval=="1"):
         try:
             update_bitSkins_base_to_excel(game_code)
         except:
             print("error")
             continue
     elif(approval!="2"):
         print("Wrong number")
         continue
     print("Update steam price?")
     print("1 - Yes")
     print("2 - No")
     approval = input()
     if(approval=="1"):
         print("Shoose method:")
         print("1 - Accurate and slow")
         print("2 - Inaccurate and fast")
         approval=input()
         if(approval=="1"):
             print("Use Tor-proxy?")
             print("1 - Yes")
             print("2 - No")
             approval=input()
             if(approval=="1"):
                 use_tor_proxy=True
             elif(approval=="2"):
                 use_tor_proxy=False
             else:
                 print("Wrong number")
                 continue
             update_accurate_price_steam_to_excel(game_code,use_tor_proxy,DEFAULT_TOR_CHANNEL)
         elif(approval=="2"):
             get_fast_price_steam_api(game_code)
         else:
             print("Wrong number")
             continue
     elif(approval!="2"):
         print("Wrong number")
         continue
    


program()